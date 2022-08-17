import requests
from bs4 import BeautifulSoup
import openpyxl


# 1번 문제
response_1 = requests.get('https://kin.naver.com/search/list.naver?query=%ED%8C%8C%EC%9D%B4%EC%8D%AC')
html_1 = response_1.text
soup_1 = BeautifulSoup(html_1, 'html.parser')

search_area_1 = soup_1.select_one('dl')

title_1 = search_area_1.select_one('dt').text.strip()
link_1 = search_area_1.select_one('a').attrs['href']
date_1 = search_area_1.select_one('.txt_inline').text
contents_1 = search_area_1.select('dd')[1].text

# print(title_1,link_1,date_1,contents_1,sep='\n')


# 2번 문제
search_area_2 = soup_1.select('dl')

title_2 =[]
link_2 =[]
date_2 = []
contents_2 =[]

for i in search_area_2:
    title_2.append(i.select_one('dt').text.strip())
    link_2.append(i.select_one('a').attrs['href'])
    date_2.append(i.select_one('.txt_inline').text)
    contents_2.append(i.select('dd')[1].text)
    
# print(title_2,date_2,contents_2,sep='\n')


# # 3번 문제
# search_page_1 = input('몇 페이지까지 검색할까요?')

# title_3 =[]
# link_3 =[]
# date_3 = []
# contents_3 =[]

# for i in range(1,int(search_page_1)+1):
#     response_2 = requests.get(f'https://kin.naver.com/search/list.naver?query=%ED%8C%8C%EC%9D%B4%EC%8D%AC&page={i}')
#     html_2 = response_2.text
#     soup_2 = BeautifulSoup(html_2, 'html.parser')

#     search_area_3 = soup_2.select('dl')

#     for i in search_area_3:
#         title_3.append(i.select_one('dt').text.strip())
#         link_3.append(i.select_one('a').attrs['href'])
#         date_3.append(i.select_one('.txt_inline').text)
#         contents_3.append(i.select('dd')[1].text)


# # 4번 문제
# search_keyword = input('검색할 키워드를 입력하세요')
# search_page_2 = input('몇 페이지까지 검색할까요?')

# title_4 =[]
# link_4 =[]
# date_4 = []
# contents_4 =[]

# for i in range(1,int(search_page_2)+1):
#     response_3 = requests.get(f'https://kin.naver.com/search/list.naver?query={search_keyword}&page={i}')
#     html_3 = response_3.text
#     soup_3 = BeautifulSoup(html_3, 'html.parser')

#     search_area_4 = soup_3.select('dl')

#     for i in search_area_4:
#         title_4.append(i.select_one('dt').text.strip())
#         link_4.append(i.select_one('a').attrs['href'])
#         date_4.append(i.select_one('.txt_inline').text)
#         contents_4.append(i.select('dd')[1].text)


# # 5번 문제
# wb = openpyxl.Workbook()
# ws = wb.active

# ws.title = '지식IN'
# ws['A1'] = '제목'
# ws['B1'] = '날짜'
# ws['C1'] = '링크'
# ws['D1'] = '설명'

# ws.append(['1번째문제'])
# ws.append([title_1,date_1,link_1,contents_1])

# for i in range(2,5):
#     ws.append([f'{i}번째문제'])
#     for j in range(len(locals()[f'title_{i}'])):
#         ws.append([locals()[f'title_{i}'][j],locals()[f'date_{i}'][j],locals()[f'link_{i}'][j],locals()[f'contents_{i}'][j]])

# wb.save('04.requests사용법/실습2.xlsx')


# # 6번 문제
# response_in = requests.get(link_1)
# html_in = response_in.text
# soup_in = BeautifulSoup(html_in, 'html.parser')

# answer_1=[]
# for i in soup_in.select('.se-module.se-module-text'):
#     answer_1.append(i.text)

# print(answer_1)