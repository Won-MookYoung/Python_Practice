##엑셀에 저장하기
#제목 본문 (첫행 제목 본문)
#크롤링한 제목1 크롤링한본문1
#크롤링한 제목2 크롤링한본문2
'''
import requests
from bs4 import BeautifulSoup
import openpyxl

response = requests.get('https://search.naver.com/search.naver?where=news&sm=tab_jum&query=%ED%8C%8C%EC%9D%B4%EC%8D%AC')
html=response.text

soup=BeautifulSoup(html,'html.parser')

python= soup.select('.news_area')
for news in python:
    title=news.select_one('.news_tit').text
    content=news.select_one('.api_txt_lines.dsc_txt_wrap').text


wb=openpyxl.Workbook()
ws = wb.create_sheet('python 기사')
ws['A1']= '제목'     
ws['B1']= '본문'

print(title,content)

'''

import requests
from bs4 import BeautifulSoup
import openpyxl

wb=openpyxl.Workbook()
ws = wb.create_sheet('python 기사')
ws['A1']= '제목'     
ws['B1']= '본문'

response = requests.get('https://search.naver.com/search.naver?where=news&sm=tab_jum&query=%ED%8C%8C%EC%9D%B4%EC%8D%AC')
html=response.text

soup=BeautifulSoup(html,'html.parser')

python= soup.select('.news_area')
i=1
for news in python:
    title=news.select_one('.news_tit').text
    content=news.select_one('.api_txt_lines.dsc_txt_wrap').text

    print(i,title,content,sep=' \n')
    i+=1


    ws.append([title,content])

wb.save('python사용법.xlsx')