import openpyxl

#1. 엑셀 파일생성
#Workbook은 클래스
# 만들었지만 아직 저장한건아님
wb=openpyxl.Workbook()

#2. 워크시트 생성
ws = wb.create_sheet('오징어게임')

#3. 데이터 추가 
ws['A1']= '참가번호'     #A1 셀
ws['B1']= '성명'

ws['A2']= 1
ws['B2']= '오일남'

wb.save('참가자_data.xlsx')

#4. 한 행 추가
ws.append([456,'성기훈'])
wb.save('참가자_data.xlsx')


wb.save('04. request사용법/참가자_data.xlsx')  #다른경로에 저장